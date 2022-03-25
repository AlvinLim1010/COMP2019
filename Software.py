import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import *
import mysql.connector
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
import xlsxwriter
from PIL import Image, ImageTk
from datetime import datetime

# Sklearn
from sklearn.svm import SVR  # for building SVR model
from sklearn.multioutput import MultiOutputRegressor  # To make the output be independent to each other

# Data Manipulation
import pandas as pd  # for data manipulation
import numpy as np  # for data manipulation

TitleFont = ("Arial", 35)


class tkinterApp(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        self.resizable(0, 0)
        windowWidth = self.winfo_screenwidth()
        windowHeight = self.winfo_screenheight()
        appWidth = 1100
        appHeight = 700

        x = int((windowWidth / 2) - (appWidth / 2))
        y = int((windowHeight / 2) - (appHeight / 2))

        self.geometry("{}x{}+{}+{}".format(appWidth, appHeight, x, y))

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (HomePage, PredictionPage, PredictionPage2, HistoryPage, OutputPage):
            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(PredictionPage)

        # Get the data from excel to train the AI model
        global updated_df
        dataset = pd.read_excel("Excel Data File/Data_SEGP.xlsx")
        dataset.drop("Unnamed: 18", axis=1, inplace=True)
        dataset.drop("Unnamed: 19", axis=1, inplace=True)
        dataset.drop("Unnamed: 20", axis=1, inplace=True)
        dataset.drop("Unnamed: 21", axis=1, inplace=True)
        updated_df = dataset
        updated_df['CH4'] = updated_df['CH4'].fillna(updated_df['CH4'].median()).round(1)
        updated_df['Volume Biogas/L'] = updated_df['Volume Biogas/L'].fillna(
        updated_df['Volume Biogas/L'].median()).round(1)
        updated_df['H2S'] = updated_df['H2S'].fillna(updated_df['H2S'].median()).round(1)
        updated_df['CO2'] = updated_df['CO2'].fillna(updated_df['CO2'].median()).round(1)
        updated_df['pH Reactor'] = updated_df['pH Reactor'].fillna(updated_df['pH Reactor'].median()).round(1)
        updated_df['pH F'] = updated_df['pH F'].fillna(updated_df['pH F'].median()).round(1)
        updated_df['pH Eff'] = updated_df['pH Eff'].fillna(updated_df['pH Eff'].median()).round(1)
        updated_df['ORL gCOD/Ld'] = updated_df['ORL gCOD/Ld'].fillna(updated_df['ORL gCOD/Ld'].median()).round(1)
        updated_df['COD F/mg/L'] = updated_df['COD F/mg/L'].fillna(updated_df['COD F/mg/L'].median()).round(1)
        updated_df['COD Eff/mg/L'] = updated_df['COD Eff/mg/L'].fillna(updated_df['COD Eff/mg/L'].median()).round(1)
        updated_df['COD removal'] = updated_df['COD removal'].fillna(updated_df['COD removal'].median()).round(1)
        updated_df['BOD F/mg/L'] = updated_df['BOD F/mg/L'].fillna(updated_df['BOD F/mg/L'].median()).round(1)
        updated_df['BOD Eff/mg/L'] = updated_df['BOD Eff/mg/L'].fillna(updated_df['BOD Eff/mg/L'].median()).round(1)
        updated_df['BOD removal'] = updated_df['BOD removal'].fillna(updated_df['BOD removal'].median()).round(1)
        updated_df['TSS F/mg/L'] = updated_df['TSS F/mg/L'].fillna(updated_df['TSS F/mg/L'].median()).round(1)
        updated_df['TSS Eff/mg/L'] = updated_df['TSS Eff/mg/L'].fillna(updated_df['TSS Eff/mg/L'].median()).round(1)
        updated_df['TSS removal'] = updated_df['TSS removal'].fillna(updated_df['TSS removal'].median()).round(1)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()
        self.title('POMEPAI')


class HomePage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        # Background image
        bg = Image.open("Pictures/GUI1.png")
        resizebg = bg.resize((1100, 700), Image.ANTIALIAS)

        self.newbg = ImageTk.PhotoImage(resizebg)
        rectangle = tk.Canvas(self, width=1100, height=700, borderwidth=0, highlightthickness=0)
        rectangle.create_image(0, 0, image=self.newbg, anchor=NW)
        rectangle.create_text(130, 155, text="POME Prediction AI", fill="#DDAA85", font=('Raleway', 17, 'bold'))
        rectangle.grid(row=0, column=0)

        aboutButton = tk.Button(self, text="ABOUT", fg="#DDAA85", bg="#4F3D2F", width=30, height=2, bd=0,
                                activebackground="#4F3D2F",
                                activeforeground="#DDAA85", font=('Raleway', 10, 'bold'),
                                command=lambda: controller.show_frame(HomePage))
        aboutButton.grid(column=0, row=0, padx=(0, 840), pady=(0, 135))

        predictButton = tk.Button(self, text="PREDICTION (SINGLE)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  activebackground="Black",
                                  activeforeground="#DDAA85", font=('Raleway', 10, 'bold'),
                                  command=lambda: controller.show_frame(PredictionPage))
        predictButton.grid(column=0, row=0, padx=(0, 840), pady=(100, 70))

        predictButton2 = tk.Button(self, text="PREDICTION (EXCEL)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                   activebackground="Black", activeforeground="#DDAA85",
                                   font=('Raleway', 10, 'bold'),
                                   command=lambda: controller.show_frame(PredictionPage2))
        predictButton2.grid(column=0, row=0, padx=(0, 840), pady=(300, 100))

        historyButton = tk.Button(self, text="HISTORY", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  activebackground="#4F3D2F", activeforeground="BLACK",
                                  font=('Raleway', 10, 'bold'),
                                  command=lambda: controller.show_frame(HistoryPage))
        historyButton.grid(column=0, row=0, padx=(0, 840), pady=(460, 100))


class PredictionPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        # Background image
        bg = Image.open("Pictures/GUI2.png")
        resizebg = bg.resize((1100, 700), Image.ANTIALIAS)

        self.newbg = ImageTk.PhotoImage(resizebg)
        rectangle = tk.Canvas(self, width=1100, height=700, borderwidth=0, highlightthickness=0)
        rectangle.create_image(0, 0, image=self.newbg, anchor=NW)
        rectangle.create_text(130, 155, text="POME Prediction AI", fill="#DDAA85", font=('Raleway', 17, 'bold'))
        rectangle.create_text(680, 100, text="Prediction For Single Point", fill="#DDAA85",
                              font=('Raleway', 17, 'bold'))
        rectangle.create_text(675, 450, text="Select Biogas For Prediction", fill="#4F3D2F",
                              font=('Raleway', 10, 'bold'))
        rectangle.grid(row=0, column=0)

        aboutButton = tk.Button(self, text="ABOUT", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                font=('Raleway', 10, 'bold'), activebackground="Black",
                                activeforeground="#DDAA85", command=lambda: controller.show_frame(HomePage))
        aboutButton.grid(column=0, row=0, padx=(0, 840), pady=(0, 135))

        predictButton = tk.Button(self, text="PREDICTION (SINGLE)", fg="#DDAA85", bg="#4F3D2F", width=30, height=2,
                                  bd=0, font=('Raleway', 10, 'bold'), activebackground="#4F3D2F",
                                  activeforeground="#DDAA85", command=lambda: controller.show_frame(PredictionPage))
        predictButton.grid(column=0, row=0, padx=(0, 840), pady=(100, 70))

        predictButton2 = tk.Button(self, text="PREDICTION (EXCEL)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                   activebackground="Black", activeforeground="#DDAA85",
                                   font=('Raleway', 10, 'bold'),
                                   command=lambda: controller.show_frame(PredictionPage2))
        predictButton2.grid(column=0, row=0, padx=(0, 840), pady=(300, 100))

        historyButton = tk.Button(self, text="HISTORY", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  font=('Raleway', 10, 'bold'), activebackground="Black",
                                  activeforeground="#DDAA85", command=lambda: controller.show_frame(HistoryPage))
        historyButton.grid(column=0, row=0, padx=(0, 840), pady=(460, 100))

        input1 = tk.StringVar()
        input2 = tk.StringVar()
        input3 = tk.StringVar()
        input4 = tk.StringVar()
        input5 = tk.StringVar()
        input6 = tk.StringVar()
        input7 = tk.StringVar()
        input8 = tk.StringVar()
        input9 = tk.StringVar()

        input1Entered = tk.Entry(self, width=30, textvariable=input1)
        input1Entered.grid(column=0, row=0, padx=(0, 250), pady=(0, 250))

        input2Entered = tk.Entry(self, width=30, textvariable=input2)
        input2Entered.grid(column=0, row=0, padx=(0, 250), pady=(0, 80))

        input3Entered = tk.Entry(self, width=30, textvariable=input3)
        input3Entered.grid(column=0, row=0, padx=(0, 250), pady=(90, 0))

        input4Entered = tk.Entry(self, width=30, textvariable=input4)
        input4Entered.grid(column=0, row=0, padx=(260, 0), pady=(0, 250))

        input5Entered = tk.Entry(self, width=30, textvariable=input5)
        input5Entered.grid(column=0, row=0, padx=(260, 0), pady=(0, 80))

        input6Entered = tk.Entry(self, width=30, textvariable=input6)
        input6Entered.grid(column=0, row=0, padx=(260, 0), pady=(90, 0))

        input7Entered = tk.Entry(self, width=30, textvariable=input7)
        input7Entered.grid(column=0, row=0, padx=(780, 0), pady=(0, 250))

        input8Entered = tk.Entry(self, width=30, textvariable=input8)
        input8Entered.grid(column=0, row=0, padx=(780, 0), pady=(0, 80))

        input9Entered = tk.Entry(self, width=30, textvariable=input9)
        input9Entered.grid(column=0, row=0, padx=(780, 0), pady=(90, 0))

        CH4Check = tk.IntVar()
        CO2Check = tk.IntVar()
        H2SCheck = tk.IntVar()
        tk.Checkbutton(self, text='CH4', variable=CH4Check, bg="#E4BC9E", activebackground="#E4BC9E").grid(column=0,
                                                                                                           row=0, padx=(150, 0), pady=(300, 0))
        tk.Checkbutton(self, text='CO2', variable=CO2Check, bg="#E4BC9E", activebackground="#E4BC9E").grid(column=0,
                                                                                                           row=0, padx=(250, 0), pady=(300, 0))
        tk.Checkbutton(self, text='H2S', variable=H2SCheck, bg="#E4BC9E", activebackground="#E4BC9E").grid(column=0,
                                                                                                           row=0, padx= (350, 0), pady=(300, 0))

        def validateCheckBox():
            if CO2Check.get() == 0 and CH4Check.get() == 0 and H2SCheck.get() == 0:
                file_label = tk.Label(self, text='Invalid input', foreground='red')
                file_label.grid(column=0, row=0, padx=(350, 80), pady=(580, 20))
                file_label.after(3000, lambda: file_label.destroy())
            else:
                predictSingleOutput(input1.get(), input4.get(), input7.get(),
                                    input2.get(), input5.get(), input8.get(),
                                    input3.get(), input6.get(), input9.get(),
                                    CH4Check.get(), CO2Check.get(),
                                    H2SCheck.get())

        def nonInteger():
            # if (input1.get().isnumeric() or input2.get().isnumeric() or input3.get().isnumeric() or input4.get().isnumeric() or input5.get().isnumeric() or input6.get().isnumeric() or input7.get().isnumeric() or input8.get().isnumeric() or input9.get().isnumeric()):
            #     validateCheckBox()

            if input1.get() == '' and input2.get() == '' and input3.get() == '' and input4.get() == '' and input5.get() == '' and input6.get() == '' and input7.get() == '' and input8.get() == '' and input9.get() == '':
                if input1.get().isalpha() or input2.get().isalpha() or input3.get().isalpha() or input4.get().isalpha() or input5.get().isalpha() or input6.get().isalpha() or input7.get().isalpha() or input8.get().isalpha() or input9.get().isalpha():
                    file_label = tk.Label(self, text='Invalid input', foreground='red')
                    file_label.grid(column=0, row=0, padx=(350, 80), pady=(580, 20))
                    file_label.after(3000, lambda: file_label.destroy())

            else:
                validateCheckBox()

        doPredictionButton = tk.Button(self, text="DO PREDICTION", fg="#4F3D2F", bg="white", width=23, height=2, bd=0,
                                       command=lambda: nonInteger())
        doPredictionButton.grid(column=0, row=0, padx=(340, 80), pady=(520, 20))

        clearButton = tk.Button(self, text="CLEAR ALL", fg="#4F3D2F", bg="white", width=13, height=1, bd=0,
                                command=lambda: clear())
        clearButton.grid(column=0, row=0, padx=(850, 80), pady=(520, 20))

        def clear():
            input1.set("")
            input2.set("")
            input3.set("")
            input4.set("")
            input5.set("")
            input6.set("")
            input7.set("")
            input8.set("")
            input9.set("")
            CH4Check.set(0)
            CO2Check.set(0)
            H2SCheck.set(0)

        def predictSingleOutput(ph_input1, ph_input2, ph_input3, cod_input1, cod_input2, cod_input3, bod_input1,
                                bod_input2, bod_input3, ch4_input, co2_input, h2s_input):
            xInput = [[]]
            x = []
            y = []
            xInput = np.asarray(xInput)
            x = np.array(x)
            y = np.array(y)

            if ph_input1 != '':
                xInput = np.concatenate((xInput, [[float(ph_input1)]]), 1)
                x = np.concatenate((x, ['pH Reactor']))
            if ph_input2 != '':
                xInput = np.concatenate((xInput, [[float(ph_input2)]]), 1)
                x = np.concatenate((x, ['pH F']))
            if ph_input3 != '':
                xInput = np.concatenate((xInput, [[float(ph_input3)]]), 1)
                x = np.concatenate((x, ['pH Eff']))
            if cod_input1 != '':
                xInput = np.concatenate((xInput, [[float(cod_input1)]]), 1)
                x = np.concatenate((x, ['COD F/mg/L']))
            if cod_input2 != '':
                xInput = np.concatenate((xInput, [[float(cod_input2)]]), 1)
                x = np.concatenate((x, ['COD Eff/mg/L']))
            if cod_input3 != '':
                xInput = np.concatenate((xInput, [[float(cod_input3)]]), 1)
                x = np.concatenate((x, ['COD removal']))
            if bod_input1 != '':
                xInput = np.concatenate((xInput, [[float(bod_input1)]]), 1)
                x = np.concatenate((x, ['BOD F/mg/L']))
            if bod_input2 != '':
                xInput = np.concatenate((xInput, [[float(bod_input2)]]), 1)
                x = np.concatenate((x, ['BOD Eff/mg/L']))
            if bod_input3 != '':
                xInput = np.concatenate((xInput, [[float(bod_input3)]]), 1)
                x = np.concatenate((x, ['BOD removal']))

            if ch4_input == 1:
                y = np.concatenate((y, ['CH4']))
            if co2_input == 1:
                y = np.concatenate((y, ['CO2']))
            if h2s_input == 1:
                y = np.concatenate((y, ['H2S']))

            # Getting dataset to fit to Model
            xData = updated_df[x].values
            yData = updated_df[y].values

            # AI Model
            svr = SVR(kernel='rbf')
            multiOutputSVR = MultiOutputRegressor(svr)
            multiOutputSVR = multiOutputSVR.fit(xData, yData)

            y_pred = multiOutputSVR.predict(xInput)
            y_pred = np.around(y_pred, 3)

            dataHeadings = []
            dataXAndY = [[]]
            dataHeadings = np.array(dataHeadings)
            dataXAndY = np.array(dataXAndY)

            dataHeadings = np.concatenate((dataHeadings, y))
            dataHeadings = np.concatenate((dataHeadings, x))
            dataHeadingOnScreen = dataHeadings.tolist()
            dataXAndY = np.concatenate((dataXAndY, y_pred), 1)
            dataXAndY = np.concatenate((dataXAndY, xInput), 1)
            dataOnScreen = dataXAndY.tolist()
            dataHeadings = np.vstack((dataHeadings, dataXAndY))

            clear()

            def outputScreenSinglePrediction(headingsToDisplayOnScreen, dataToDisplayInFile, dataToDisplayOnScreen):
                outputDisplaying = Tk()
                outputDisplaying.title("OutputScreen")
                outputDisplaying.resizable(0, 0)

                def downloadExcelOutput():
                    dateAndTime = datetime.now()
                    dateAndTime = dateAndTime.strftime("%d-%m-%Y_%H.%M.%S")

                    filename = "OutputFiles/Output_" + dateAndTime + ".xlsx"
                    workbook = xlsxwriter.Workbook(filename)
                    worksheet = workbook.add_worksheet("Output")

                    col = 0

                    for row, dataInFile in enumerate(dataToDisplayInFile):
                        worksheet.write_row(row, col, dataInFile)

                    worksheet.set_column(0, (len(dataToDisplayInFile[0]) - 1), 12)

                    workbook.close()

                tree_frame = Frame(outputDisplaying)
                tree_frame.pack()

                tree_scroll = Scrollbar(tree_frame)
                tree_scroll.pack(side=RIGHT, fill=Y)

                outputTree = ttk.Treeview(tree_frame, show='headings', height=8, yscrollcommand=tree_scroll.set,
                                          selectmode="none")
                outputTree.pack()

                tree_scroll.config(command=outputTree.yview)

                outputTree['columns'] = headingsToDisplayOnScreen
                outputTree.column("#0", width=0, stretch=False)
                length = int(800 / len(headingsToDisplayOnScreen))

                for i in headingsToDisplayOnScreen:
                    outputTree.column(i, anchor=CENTER, width=length, stretch=False)
                    outputTree.heading(i, text=i, anchor=CENTER)

                count = 0
                for data in dataToDisplayOnScreen:
                    outputTree.insert(parent='', index='end', iid=count, text="", values=data)
                    count += 1

                ttk.Button(outputDisplaying, text='Download Output', command=lambda: downloadExcelOutput()).pack()

                outputDisplaying.mainloop()

            outputScreenSinglePrediction(dataHeadingOnScreen, dataHeadings, dataOnScreen)


class PredictionPage2(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        listbox = Listbox(self, height=144, width=144, selectmode="multiple")
        listbox.pack

        def getExcel():
            filename = filedialog.askopenfilename(title="Open A File", filetypes=[('Excel Files', '*.xlsx')])
            e1.insert(0, filename)
            e1.config(text=filename)

            global df
            df = pd.read_excel(filename)

            box1.delete(0, END)

            for item in df.columns:
                box1.insert(END, item)

            file_label = tk.Label(self, text='File Uploaded Successfully!', foreground='green')
            file_label.grid(column=0, row=0, padx=(350, 80), pady=(580, 20))
            file_label.after(3000, lambda: file_label.destroy())

        # Background image
        bg = Image.open("Pictures/GUI3.png")
        resizebg = bg.resize((1100, 700), Image.ANTIALIAS)

        self.newbg = ImageTk.PhotoImage(resizebg)
        rectangle = tk.Canvas(self, width=1100, height=700, borderwidth=0, highlightthickness=0)
        rectangle.create_image(0, 0, image=self.newbg, anchor=NW)
        rectangle.create_text(130, 155, text="POME Prediction AI", fill="#DDAA85", font=('Raleway', 17, 'bold'))
        rectangle.create_text(680, 100, text="Prediction For Excel File", fill="#DDAA85", font=('Raleway', 17, 'bold'))
        rectangle.create_text(835, 190, text="(ONLY .xlsx FILES)", fill="Black", font=('Raleway', 10))
        rectangle.grid(row=0, column=0)

        aboutButton = tk.Button(self, text="ABOUT", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                activebackground="Black", activeforeground="#DDAA85",
                                font=('Raleway', 10, 'bold'),
                                command=lambda: controller.show_frame(HomePage))
        aboutButton.grid(column=0, row=0, padx=(0, 840), pady=(0, 135))

        predictButton = tk.Button(self, text="PREDICTION (SINGLE)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  activebackground="Black", activeforeground="#DDAA85",
                                  font=('Raleway', 10, 'bold'),
                                  command=lambda: controller.show_frame(PredictionPage))
        predictButton.grid(column=0, row=0, padx=(0, 840), pady=(100, 70))

        predictButton2 = tk.Button(self, text="PREDICTION (EXCEL)", fg="#DDAA85", bg="#4F3D2F", width=30, height=2,
                                   bd=0,
                                   activebackground="#4F3D2F", activeforeground="#DDAA85",
                                   font=('Raleway', 10, 'bold'),
                                   command=lambda: controller.show_frame(PredictionPage2))
        predictButton2.grid(column=0, row=0, padx=(0, 840), pady=(300, 100))

        historyButton = tk.Button(self, text="HISTORY", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  activebackground="Black", activeforeground="#DDAA85",
                                  font=('Raleway', 10, 'bold'),
                                  command=lambda: controller.show_frame(HistoryPage))
        historyButton.grid(column=0, row=0, padx=(0, 840), pady=(460, 100))

        def remove():
            box2.delete(0, END)
            box3.delete(0, END)

        def getX():
            global x_v
            x_v = []
            s = box1.curselection()
            box2.delete(0, END)
            for i in s:
                data = box1.get(i)
                x_v.append(data)

            for data in x_v:
                box2.insert(END, data)

            box1.select_clear(0, END)

        def getY():
            global y_v
            y_v = []
            s = box1.curselection()
            box3.delete(0, END)
            for i in s:
                data = box1.get(i)
                y_v.append(data)

            for data in y_v:
                box3.insert(END, data)

            box1.select_clear(0, END)

        def predictFileOutput():
            # Getting dataset to fit to Model
            xData = updated_df[x_v].values
            yData = updated_df[y_v].values

            # Getting dataset to do prediction
            xInput = df[x_v].values

            # AI Model
            svr = SVR(kernel='rbf')
            multiOutputSVR = MultiOutputRegressor(svr)
            multiOutputSVR = multiOutputSVR.fit(xData, yData)

            y_pred = multiOutputSVR.predict(xInput)
            y_pred = np.around(y_pred, 3)

            fullData = np.concatenate((y_pred, xInput), axis=1)
            heading = np.concatenate((y_v, x_v))
            dataFiles = np.vstack((heading, fullData))
            headingScreen = heading.tolist()
            fullData = fullData.tolist()

            def outputScreenSinglePrediction(headingsToDisplayOnScreen, dataToDisplayInFile, dataToDisplayOnScreen):
                outputDisplaying = Tk()
                outputDisplaying.title("OutputScreen")
                outputDisplaying.resizable(0, 0)

                def downloadExcelOutput():
                    dateAndTime = datetime.now()
                    dateAndTime = dateAndTime.strftime("%d-%m-%Y_%H.%M.%S")

                    filename = "OutputFiles/Output_" + dateAndTime + ".xlsx"
                    workbook = xlsxwriter.Workbook(filename)
                    worksheet = workbook.add_worksheet("Output")

                    col = 0

                    for row, dataInFile in enumerate(dataToDisplayInFile):
                        worksheet.write_row(row, col, dataInFile)

                    worksheet.set_column(0, (len(dataToDisplayInFile[0]) - 1), 12)

                    workbook.close()

                tree_frame = Frame(outputDisplaying)
                tree_frame.pack()

                tree_scroll = Scrollbar(tree_frame)
                tree_scroll.pack(side=RIGHT, fill=Y)

                outputTree = ttk.Treeview(tree_frame, show='headings', height=8, yscrollcommand=tree_scroll.set,
                                          selectmode="none")
                outputTree.pack()

                tree_scroll.config(command=outputTree.yview)

                outputTree['columns'] = headingsToDisplayOnScreen
                outputTree.column("#0", width=0, stretch=False)
                length = int(800 / len(headingsToDisplayOnScreen))

                for i in headingsToDisplayOnScreen:
                    outputTree.column(i, anchor=CENTER, width=length, stretch=False)
                    outputTree.heading(i, text=i, anchor=CENTER)

                count = 0
                for data in dataToDisplayOnScreen:
                    outputTree.insert(parent='', index='end', iid=count, text="", values=data)
                    count += 1

                ttk.Button(outputDisplaying, text='Download Output', command=lambda: downloadExcelOutput()).pack()

                outputDisplaying.mainloop()

            outputScreenSinglePrediction(headingScreen, dataFiles, fullData)

        Button(self, text='Select X', activeforeground="white", activebackground="black", bd=0,
               command=lambda: getX()).grid(column=0, row=0, padx=(157, 0), pady=(300, 0))
        Button(self, text='Select Y', activeforeground="white", activebackground="black", bd=0,
               command=lambda: getY()).grid(column=0, row=0, padx=(370, 0), pady=(300, 0))
        Button(self, text='Remove', activeforeground="white", activebackground="black", bd=0,
               command=lambda: remove()).grid(column=0, row=0, padx=(400, 0), pady=(400, 0))

        e1 = Entry(self, text='')
        e1.grid(column=0, row=0, padx=(260, 0), pady=(0, 270))

        box1 = Listbox(self, height=15, width=25, selectmode='multiple', activestyle='none')
        box1.grid(row=0, column=0, padx=(0, 50), pady=(15, 0))

        box2 = Listbox(self, height=15, width=25)
        box2.grid(row=0, column=0, padx=(263, 0), pady=(15, 0))

        box3 = Listbox(self, height=15, width=25)
        box3.grid(row=0, column=0, padx=(575, 0), pady=(15, 0))

        doPredictionButton = tk.Button(self, text="DO PREDICTION", fg="#4F3D2F", bg="white", width=23, height=2, bd=0,
                                       command=lambda: predictFileOutput())
        doPredictionButton.grid(column=0, row=0, padx=(340, 80), pady=(520, 20))

        importButton = tk.Button(self, text="IMPORT FILE", command=getExcel, fg="#4F3D2F", bg="white", width=10,
                                 height=1, bd=0)
        importButton.grid(column=0, row=0, padx=(575, 0), pady=(0, 270))


class HistoryPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        # Background image
        bg = Image.open("Pictures/GUI4.png")
        resizebg = bg.resize((1100, 700), Image.ANTIALIAS)

        self.newbg = ImageTk.PhotoImage(resizebg)
        rectangle = tk.Canvas(self, width=1100, height=700, borderwidth=0, highlightthickness=0)
        rectangle.create_image(0, 0, image=self.newbg, anchor=NW)
        rectangle.create_text(130, 155, text="POME Prediction AI", fill="#DDAA85", font=('Raleway', 17, 'bold'))
        rectangle.grid(row=0, column=0)

        aboutButton = tk.Button(self, text="ABOUT", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                activebackground="Black", activeforeground="#DDAA85",
                                font=('Raleway', 10, 'bold'),
                                command=lambda: controller.show_frame(HomePage))
        aboutButton.grid(column=0, row=0, padx=(0, 840), pady=(0, 135))

        predictButton = tk.Button(self, text="PREDICTION (SINGLE)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  activebackground="Black", activeforeground="#DDAA85",
                                  font=('Raleway', 10, 'bold'),
                                  command=lambda: controller.show_frame(PredictionPage))
        predictButton.grid(column=0, row=0, padx=(0, 840), pady=(100, 70))

        predictButton2 = tk.Button(self, text="PREDICTION (EXCEL)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                   activebackground="Black", activeforeground="#DDAA85",
                                   font=('Raleway', 10, 'bold'),
                                   command=lambda: controller.show_frame(PredictionPage2))
        predictButton2.grid(column=0, row=0, padx=(0, 840), pady=(300, 100))

        historyButton = tk.Button(self, text="HISTORY", fg="#DDAA85", bg="#4F3D2F", width=30, height=2, bd=0,
                                  activebackground="#4F3D2F", activeforeground="#DDAA85",
                                  font=('Raleway', 10, 'bold'),
                                  command=lambda: controller.show_frame(HistoryPage))
        historyButton.grid(column=0, row=0, padx=(0, 840), pady=(460, 100))


class OutputPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        # Background image
        bg = Image.open("Pictures/GUI5.png")
        resizebg = bg.resize((1100, 700), Image.ANTIALIAS)

        self.newbg = ImageTk.PhotoImage(resizebg)
        rectangle = tk.Canvas(self, width=1100, height=700, borderwidth=0, highlightthickness=0)
        rectangle.create_image(0, 0, image=self.newbg, anchor=NW)
        rectangle.create_text(195, 745, text="(ONLY .xlsx FILES)", fill="White", font=('Raleway', 10))
        rectangle.create_text(130, 155, text="POME Prediction AI", fill="#DDAA85", font=('Raleway', 17, 'bold'))
        rectangle.grid(row=0, column=0)

        aboutButton = tk.Button(self, text="ABOUT", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                font=('Raleway', 10, 'bold'), activebackground="Black",
                                activeforeground="#DDAA85", command=lambda: controller.show_frame(HomePage))
        aboutButton.grid(column=0, row=0, padx=(0, 840), pady=(0, 135))

        predictButton = tk.Button(self, text="PREDICTION (SINGLE)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  font=('Raleway', 10, 'bold'), activebackground="#4F3D2F",
                                  activeforeground="#DDAA85", command=lambda: controller.show_frame(PredictionPage))
        predictButton.grid(column=0, row=0, padx=(0, 840), pady=(100, 70))

        predictButton2 = tk.Button(self, text="PREDICTION (EXCEL)", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                   activebackground="Black", activeforeground="#DDAA85",
                                   font=('Raleway', 10, 'bold'),
                                   command=lambda: controller.show_frame(PredictionPage2))
        predictButton2.grid(column=0, row=0, padx=(0, 840), pady=(300, 100))

        historyButton = tk.Button(self, text="HISTORY", fg="#DDAA85", bg="Black", width=30, height=2, bd=0,
                                  activebackground="#4F3D2F", activeforeground="BLACK",
                                  font=('Raleway', 10, 'bold'),
                                  command=lambda: controller.show_frame(HistoryPage))
        historyButton.grid(column=0, row=0, padx=(0, 840), pady=(460, 100))

        newDataButton = tk.Button(self, text="INPUT NEW DATA", fg="#4F3D2F", bg="white", width=17, height=1, bd=0,
                                  command=lambda: controller.show_frame(PredictionPage))
        newDataButton.grid(column=0, row=0, padx=(10, 380), pady=(520, 20))

        downloadButton = tk.Button(self, text="DOWNLOAD OUTPUT", fg="#4F3D2F",
                                   bg="white", width=17, height=1, bd=0)
        downloadButton.grid(column=0, row=0, padx=(350, 80), pady=(520, 20))


app = tkinterApp()
app.mainloop()
